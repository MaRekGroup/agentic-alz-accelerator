"""
Excel Reporter — generates branded Excel workbooks from WARA assessment results.

Produces a seven-sheet workbook for WARA reporting:
- Dashboard
- Recommendations
- Findings Detail
- Resource Inventory
- Pillar Scores
- Remediation Roadmap
- Summary Statistics
"""

import logging
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference, Series
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    Workbook = None
    BarChart = None
    Reference = None
    Series = None
    Alignment = None
    Font = None
    PatternFill = None
    get_column_letter = None
    _OPENPYXL_IMPORT_ERROR: ImportError | None = exc
else:
    _OPENPYXL_IMPORT_ERROR = None

from src.tools.discovery import DiscoveryResult
from src.tools.wara_engine import AssessmentResult, Finding, PillarScore

logger = logging.getLogger(__name__)

_BRAND_BLUE = "0078D4"
_BRAND_DARK = "242424"
_BRAND_LIGHT_GRAY = "F2F2F2"
_HEADER_FONT = "Segoe UI"

_PILLAR_NAMES = {
    "security": "Security",
    "reliability": "Reliability",
    "cost_optimization": "Cost Optimization",
    "operational_excellence": "Operational Excellence",
    "performance": "Performance Efficiency",
}

_PILLAR_ORDER = [
    "security",
    "reliability",
    "cost_optimization",
    "operational_excellence",
    "performance",
]

_SEVERITY_ORDER = {
    "critical": 0,
    "high": 1,
    "medium": 2,
    "low": 3,
}

_SEVERITY_FILL_COLORS = {
    "critical": "F4CCCC",
    "high": "FCE5CD",
    "medium": "FFF2CC",
    "low": "D9EAF7",
}

_SCORE_FILL_GOOD = "C6EFCE"
_SCORE_FILL_WARNING = "FFEB9C"
_SCORE_FILL_RISK = "F4CCCC"


class ExcelReporter:
    """Generates Excel action plan workbooks for WARA assessment results."""

    def __init__(self, output_dir: Path):
        """Initialize the reporter.

        Args:
            output_dir: Directory where the workbook will be written.
        """
        if _OPENPYXL_IMPORT_ERROR is not None:
            raise ImportError("openpyxl is required for Excel report generation") from _OPENPYXL_IMPORT_ERROR

        self.output_dir = Path(output_dir)

    def generate(
        self,
        discovery: DiscoveryResult,
        assessment: AssessmentResult,
        *,
        scope_label: str | None = None,
    ) -> Path:
        """Generate the Excel action plan workbook.

        Args:
            discovery: Discovery inventory used to populate resource summary data.
            assessment: Assessment results used to populate scores and findings.
            scope_label: Optional label to show in the workbook title.

        Returns:
            Path to the generated `.xlsx` workbook.
        """
        self.output_dir.mkdir(parents=True, exist_ok=True)
        label = scope_label or discovery.scope
        generated_at = datetime.now(timezone.utc)
        generated_label = self._format_timestamp(generated_at.isoformat())

        workbook = Workbook()

        dashboard_sheet = workbook.active
        dashboard_sheet.title = "Dashboard"
        self._populate_dashboard(dashboard_sheet, discovery, assessment, label, generated_label)

        recommendations_sheet = workbook.create_sheet("Recommendations")
        self._populate_recommendations(recommendations_sheet, assessment)

        findings_sheet = workbook.create_sheet("Findings Detail")
        self._populate_findings_detail(findings_sheet, assessment)

        inventory_sheet = workbook.create_sheet("Resource Inventory")
        self._populate_resource_inventory(inventory_sheet, discovery)

        pillar_scores_sheet = workbook.create_sheet("Pillar Scores")
        self._populate_pillar_scores(pillar_scores_sheet, assessment)

        roadmap_sheet = workbook.create_sheet("Remediation Roadmap")
        self._populate_remediation_roadmap(roadmap_sheet, assessment)

        summary_stats_sheet = workbook.create_sheet("Summary Statistics")
        self._populate_summary_statistics(summary_stats_sheet, discovery, assessment, generated_label)

        workbook.active = 0
        output_path = self.output_dir / "wara-action-plan.xlsx"
        workbook.save(output_path)
        logger.info("Generated Excel action plan workbook: %s", output_path)
        return output_path

    def _populate_dashboard(
        self,
        worksheet,
        discovery: DiscoveryResult,
        assessment: AssessmentResult,
        label: str,
        generated_at: str,
    ) -> None:
        worksheet.merge_cells("A1:N1")
        worksheet["A1"] = f"WARA Assessment Dashboard — {label}"
        self._style_title_cell(worksheet["A1"])

        resource_start_row = 30
        resource_end_row = self._write_dashboard_table(
            worksheet,
            start_row=resource_start_row,
            start_col=1,
            header="Resource Type",
            rows=self._resource_chart_rows(assessment.findings),
        )
        pillar_start_row = 30
        pillar_end_row = self._write_dashboard_table(
            worksheet,
            start_row=pillar_start_row,
            start_col=10,
            header="Pillar",
            rows=self._pillar_chart_rows(assessment),
        )

        resource_chart = self._build_dashboard_chart(
            worksheet,
            title="Findings by Severity per Resource Type",
            categories_col=1,
            data_start_col=2,
            header_row=resource_start_row,
            data_start_row=resource_start_row + 1,
            data_end_row=resource_end_row,
            x_axis_title="Resource Type",
        )
        pillar_chart = self._build_dashboard_chart(
            worksheet,
            title="Findings by Severity per WAF Pillar",
            categories_col=10,
            data_start_col=11,
            header_row=pillar_start_row,
            data_start_row=pillar_start_row + 1,
            data_end_row=pillar_end_row,
            x_axis_title="WAF Pillar",
        )
        worksheet.add_chart(resource_chart, "A3")
        worksheet.add_chart(pillar_chart, "J3")

        summary_row = 22
        worksheet.append([])
        worksheet[f"A{summary_row}"] = "Metric"
        worksheet[f"B{summary_row}"] = "Value"
        self._style_header_row(worksheet, summary_row)

        dashboard_stats = [
            ("Total Findings", len(assessment.findings)),
            ("Total Resources", self._resource_total(discovery)),
            ("Overall Score", round(assessment.overall_score, 1)),
            ("Date", generated_at),
        ]
        for offset, (metric, value) in enumerate(dashboard_stats, start=1):
            worksheet[f"A{summary_row + offset}"] = metric
            worksheet[f"B{summary_row + offset}"] = value
            self._style_value_cell(worksheet[f"A{summary_row + offset}"])
            self._style_value_cell(worksheet[f"B{summary_row + offset}"])

        worksheet.auto_filter.ref = f"A{summary_row}:B{summary_row + len(dashboard_stats)}"
        worksheet.freeze_panes = "A2"
        self._hide_rows(worksheet, min(resource_start_row, pillar_start_row), max(resource_end_row, pillar_end_row))
        self._auto_width(worksheet, max_row=summary_row + len(dashboard_stats))

    def _populate_recommendations(self, worksheet, assessment: AssessmentResult) -> None:
        headers = [
            "Impact",
            "Title",
            "Pillar",
            "Impacted Count",
            "Recommendation",
            "Learn More",
            "Rule ID",
        ]
        worksheet.append(headers)
        self._style_header_row(worksheet, 1)

        for recommendation in self._recommendation_rows(assessment.findings):
            worksheet.append(
                [
                    recommendation["severity"],
                    recommendation["title"],
                    recommendation["pillar"],
                    recommendation["impacted_count"],
                    recommendation["recommendation"],
                    recommendation["references"],
                    recommendation["rule_id"],
                ]
            )
            worksheet[f"A{worksheet.max_row}"].fill = self._severity_fill(recommendation["severity"])

        self._apply_wrap(worksheet, ["E", "F"])
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.freeze_panes = "A2"
        self._auto_width(worksheet)

    def _populate_findings_detail(self, worksheet, assessment: AssessmentResult) -> None:
        headers = [
            "Rule ID",
            "Title",
            "Pillar",
            "CAF Area",
            "Severity",
            "Confidence",
            "Recommendation",
            "Evidence Count",
            "References",
        ]
        worksheet.append(headers)
        self._style_header_row(worksheet, 1)

        for finding in self._sorted_findings(assessment.findings):
            worksheet.append(
                [
                    finding.rule_id,
                    finding.title,
                    _PILLAR_NAMES.get(finding.pillar, finding.pillar),
                    finding.caf_area,
                    finding.severity,
                    finding.confidence,
                    finding.recommendation,
                    len(finding.evidence),
                    "\n".join(finding.references),
                ]
            )
            worksheet[f"E{worksheet.max_row}"].fill = self._severity_fill(finding.severity)

        self._apply_wrap(worksheet, ["G", "I"])
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.freeze_panes = "A2"
        self._auto_width(worksheet)

    def _populate_resource_inventory(self, worksheet, discovery: DiscoveryResult) -> None:
        worksheet.append(["Resource Type", "Count"])
        self._style_header_row(worksheet, 1)

        for resource_type, count in self._resource_rows(discovery):
            worksheet.append([resource_type, count])

        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.freeze_panes = "A2"
        self._auto_width(worksheet)

    def _populate_pillar_scores(self, worksheet, assessment: AssessmentResult) -> None:
        headers = ["Pillar", "Score", "Critical", "High", "Medium", "Low", "Total Findings"]
        worksheet.append(headers)
        self._style_header_row(worksheet, 1)

        score_values: list[float] = []
        critical_total = 0
        high_total = 0
        medium_total = 0
        low_total = 0
        findings_total = 0

        for pillar_key in _PILLAR_ORDER:
            pillar_score = assessment.pillar_scores.get(pillar_key, PillarScore(pillar=pillar_key))
            score_values.append(pillar_score.score)
            critical_total += pillar_score.critical
            high_total += pillar_score.high
            medium_total += pillar_score.medium
            low_total += pillar_score.low
            findings_total += pillar_score.findings_count
            worksheet.append(
                [
                    _PILLAR_NAMES[pillar_key],
                    round(pillar_score.score, 1),
                    pillar_score.critical,
                    pillar_score.high,
                    pillar_score.medium,
                    pillar_score.low,
                    pillar_score.findings_count,
                ]
            )
            worksheet[f"B{worksheet.max_row}"].fill = self._score_fill(pillar_score.score)

        average_score = sum(score_values) / len(score_values) if score_values else 0.0
        worksheet.append(
            [
                "Average / Total",
                round(average_score, 1),
                critical_total,
                high_total,
                medium_total,
                low_total,
                findings_total,
            ]
        )
        total_row = worksheet.max_row
        self._style_total_row(worksheet, total_row)
        worksheet[f"B{total_row}"].fill = self._score_fill(average_score)

        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.freeze_panes = "A2"
        self._auto_width(worksheet)

    def _populate_remediation_roadmap(self, worksheet, assessment: AssessmentResult) -> None:
        headers = [
            "Priority",
            "Rule ID",
            "Title",
            "Severity",
            "Recommendation",
            "Remediation Steps",
        ]
        worksheet.append(headers)
        self._style_header_row(worksheet, 1)

        sorted_findings = self._sorted_findings(assessment.findings)
        for priority, finding in enumerate(sorted_findings, start=1):
            worksheet.append(
                [
                    priority,
                    finding.rule_id,
                    finding.title,
                    finding.severity,
                    finding.recommendation,
                    "\n".join(finding.remediation_steps),
                ]
            )
            worksheet[f"D{worksheet.max_row}"].fill = self._severity_fill(finding.severity)

        self._apply_wrap(worksheet, ["E", "F"])
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.freeze_panes = "A2"
        self._auto_width(worksheet)

    def _populate_summary_statistics(
        self,
        worksheet,
        discovery: DiscoveryResult,
        assessment: AssessmentResult,
        generated_at: str,
    ) -> None:
        worksheet.append(["Metric", "Value"])
        self._style_header_row(worksheet, 1)

        rows = [
            ("Scope", discovery.scope),
            ("Scope Type", getattr(discovery.scope_type, "value", str(discovery.scope_type))),
            ("Generated At", generated_at),
            ("Total Resources", self._resource_total(discovery)),
            ("Total Findings", len(assessment.findings)),
            ("Checks Run", assessment.checks_run),
            ("Checks Passed", assessment.checks_passed),
            ("Overall Score", round(assessment.overall_score, 1)),
        ]
        for metric, value in rows:
            worksheet.append([metric, value])
            self._style_value_cell(worksheet[f"A{worksheet.max_row}"])
            self._style_value_cell(worksheet[f"B{worksheet.max_row}"])

        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.freeze_panes = "A2"
        self._auto_width(worksheet)

    def _build_dashboard_chart(
        self,
        worksheet,
        *,
        title: str,
        categories_col: int,
        data_start_col: int,
        header_row: int,
        data_start_row: int,
        data_end_row: int,
        x_axis_title: str,
    ):
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.grouping = "clustered"
        chart.overlap = 0
        chart.gapWidth = 60
        chart.title = title
        chart.y_axis.title = "Findings"
        chart.x_axis.title = x_axis_title
        chart.height = 4.17
        chart.width = 6.25
        chart.legend.position = "r"

        categories = Reference(
            worksheet,
            min_col=categories_col,
            min_row=data_start_row,
            max_row=data_end_row,
        )
        chart.set_categories(categories)

        severity_names = list(_SEVERITY_ORDER)
        for offset, severity in enumerate(severity_names):
            values = Reference(
                worksheet,
                min_col=data_start_col + offset,
                min_row=header_row,
                max_row=data_end_row,
            )
            series = Series(values, title_from_data=True)
            chart.series.append(series)
            chart.series[-1].graphicalProperties.solidFill = _SEVERITY_FILL_COLORS[severity]
            chart.series[-1].graphicalProperties.line.solidFill = _SEVERITY_FILL_COLORS[severity]

        return chart

    def _write_dashboard_table(
        self,
        worksheet,
        *,
        start_row: int,
        start_col: int,
        header: str,
        rows: list[tuple[str, dict[str, int]]],
    ) -> int:
        headers = [header, "Critical", "High", "Medium", "Low"]
        for offset, value in enumerate(headers):
            worksheet.cell(row=start_row, column=start_col + offset, value=value)
            self._style_header_cell(worksheet.cell(row=start_row, column=start_col + offset))

        current_row = start_row + 1
        for label, counts in rows:
            worksheet.cell(row=current_row, column=start_col, value=label)
            self._style_value_cell(worksheet.cell(row=current_row, column=start_col))
            for offset, severity in enumerate(_SEVERITY_ORDER, start=1):
                worksheet.cell(row=current_row, column=start_col + offset, value=counts[severity])
                self._style_value_cell(worksheet.cell(row=current_row, column=start_col + offset))
            current_row += 1

        return current_row - 1

    def _recommendation_rows(self, findings: list[Finding]) -> list[dict[str, str | int]]:
        grouped: dict[str, list[Finding]] = defaultdict(list)
        for finding in findings:
            grouped[finding.rule_id].append(finding)

        rows: list[dict[str, str | int]] = []
        for rule_id, grouped_findings in grouped.items():
            primary = self._sorted_findings(grouped_findings)[0]
            rows.append(
                {
                    "severity": primary.severity,
                    "title": primary.title,
                    "pillar": _PILLAR_NAMES.get(primary.pillar, primary.pillar),
                    "impacted_count": self._impacted_count(grouped_findings),
                    "recommendation": primary.recommendation,
                    "references": "\n".join(self._unique_references(grouped_findings)),
                    "rule_id": rule_id,
                }
            )

        return sorted(
            rows,
            key=lambda row: (
                _SEVERITY_ORDER.get(str(row["severity"]), 99),
                -int(row["impacted_count"]),
                str(row["title"]),
            ),
        )

    def _resource_chart_rows(self, findings: list[Finding]) -> list[tuple[str, dict[str, int]]]:
        counts: dict[str, dict[str, int]] = defaultdict(lambda: dict.fromkeys(_SEVERITY_ORDER, 0))
        for finding in findings:
            for resource_type in self._finding_resource_types(finding):
                counts[resource_type][finding.severity] += 1

        if not counts:
            return [("No findings", dict.fromkeys(_SEVERITY_ORDER, 0))]

        return sorted(counts.items(), key=lambda item: (-sum(item[1].values()), item[0]))

    def _pillar_chart_rows(self, assessment: AssessmentResult) -> list[tuple[str, dict[str, int]]]:
        rows: list[tuple[str, dict[str, int]]] = []
        for pillar_key in _PILLAR_ORDER:
            pillar_score = assessment.pillar_scores.get(pillar_key, PillarScore(pillar=pillar_key))
            rows.append(
                (
                    _PILLAR_NAMES[pillar_key],
                    {
                        "critical": pillar_score.critical,
                        "high": pillar_score.high,
                        "medium": pillar_score.medium,
                        "low": pillar_score.low,
                    },
                )
            )
        return rows

    def _finding_resource_types(self, finding: Finding) -> set[str]:
        resource_types: set[str] = set()
        for evidence in finding.evidence:
            if not isinstance(evidence, dict):
                continue
            resource_type = evidence.get("type") or evidence.get("resourceType") or evidence.get("resource_type")
            if resource_type:
                resource_types.add(str(resource_type))
        if not resource_types:
            resource_types.add("Unknown")
        return resource_types

    def _impacted_count(self, findings: list[Finding]) -> int:
        impacted_resources: set[str] = set()
        for finding in findings:
            for evidence in finding.evidence:
                impacted_resources.add(repr(evidence))
        return len(impacted_resources) if impacted_resources else len(findings)

    def _unique_references(self, findings: list[Finding]) -> list[str]:
        references = {
            reference.strip()
            for finding in findings
            for reference in finding.references
            if reference and reference.strip()
        }
        return sorted(references)

    def _sorted_findings(self, findings: list[Finding]) -> list[Finding]:
        return sorted(
            findings,
            key=lambda finding: (
                _SEVERITY_ORDER.get(finding.severity, 99),
                _PILLAR_NAMES.get(finding.pillar, finding.pillar),
                finding.rule_id,
            ),
        )

    def _resource_rows(self, discovery: DiscoveryResult) -> list[tuple[str, int]]:
        by_type = discovery.resources.get("by_type", {}) if isinstance(discovery.resources, dict) else {}
        if isinstance(by_type, dict):
            rows = [(resource_type, int(count)) for resource_type, count in by_type.items()]
        else:
            rows = [
                (
                    item.get("type", "unknown"),
                    int(item.get("resource_count", item.get("count", 0))),
                )
                for item in by_type
            ]
        return sorted(rows, key=lambda row: row[1], reverse=True)

    def _resource_total(self, discovery: DiscoveryResult) -> int:
        if not isinstance(discovery.resources, dict):
            return 0
        return int(discovery.resources.get("total_count", discovery.resources.get("total", 0)))

    def _style_title_cell(self, cell) -> None:
        cell.font = Font(name=_HEADER_FONT, bold=True, size=16, color=_BRAND_DARK)
        cell.fill = PatternFill(fill_type="solid", fgColor=_BRAND_LIGHT_GRAY)
        cell.alignment = Alignment(vertical="center")

    def _style_header_row(self, worksheet, row_number: int) -> None:
        for cell in worksheet[row_number]:
            self._style_header_cell(cell)

    def _style_header_cell(self, cell) -> None:
        cell.font = Font(name=_HEADER_FONT, bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor=_BRAND_BLUE)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _style_total_row(self, worksheet, row_number: int) -> None:
        for cell in worksheet[row_number]:
            cell.font = Font(name=_HEADER_FONT, bold=True, color=_BRAND_DARK)
            cell.fill = PatternFill(fill_type="solid", fgColor=_BRAND_LIGHT_GRAY)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _style_value_cell(self, cell) -> None:
        cell.font = Font(name=_HEADER_FONT, color=_BRAND_DARK)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _apply_wrap(self, worksheet, columns: list[str]) -> None:
        for column in columns:
            for row in range(2, worksheet.max_row + 1):
                worksheet[f"{column}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
                worksheet[f"{column}{row}"].font = Font(name=_HEADER_FONT, color=_BRAND_DARK)

    def _hide_rows(self, worksheet, start_row: int, end_row: int) -> None:
        for row in range(start_row, end_row + 1):
            worksheet.row_dimensions[row].hidden = True

    def _auto_width(self, worksheet, max_row: int | None = None) -> None:
        width_row_limit = max_row or worksheet.max_row
        for column_index in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(column_index)
            max_length = 0
            for row_index in range(1, width_row_limit + 1):
                value = worksheet.cell(row=row_index, column=column_index).value
                text = "" if value is None else str(value)
                max_length = max(max_length, len(text))
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)

    def _severity_fill(self, severity: str):
        return PatternFill(
            fill_type="solid",
            fgColor=_SEVERITY_FILL_COLORS.get(severity, _SEVERITY_FILL_COLORS["low"]),
        )

    def _score_fill(self, score: float):
        if score >= 80:
            color = _SCORE_FILL_GOOD
        elif score >= 60:
            color = _SCORE_FILL_WARNING
        else:
            color = _SCORE_FILL_RISK
        return PatternFill(fill_type="solid", fgColor=color)

    def _format_timestamp(self, timestamp: str) -> str:
        try:
            value = datetime.fromisoformat(timestamp.replace("Z", "+00:00"))
        except ValueError:
            return timestamp
        return value.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
