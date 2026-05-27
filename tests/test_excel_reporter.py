"""Tests for the Excel reporter."""

import pytest
from openpyxl import load_workbook

from src.tools.discovery import DiscoveryResult, DiscoveryScope
from src.tools.excel_reporter import ExcelReporter
from src.tools.wara_engine import AssessmentResult, Finding, PillarScore


@pytest.fixture
def sample_discovery():
    return DiscoveryResult(
        scope="test-sub",
        scope_type=DiscoveryScope.SUBSCRIPTION,
        resources={
            "by_type": [
                {"type": "microsoft.compute/virtualmachines", "count": 5},
                {"type": "microsoft.storage/storageaccounts", "count": 2},
            ],
            "total_count": 7,
        },
    )


@pytest.fixture
def sample_assessment():
    finding = Finding(
        rule_id="SEC-001",
        title="TLS check",
        pillar="security",
        caf_area="security",
        alz_area="security",
        severity="high",
        confidence="high",
        recommendation="Fix TLS",
        evidence=[
            {
                "id": "/subscriptions/test-sub/resourceGroups/rg1/providers/Microsoft.Compute/virtualMachines/vm1",
                "type": "microsoft.compute/virtualmachines",
            }
        ],
        remediation_steps=["Step 1"],
        references=["https://learn.microsoft.com/example"],
    )
    result = AssessmentResult(scope="test-sub")
    result.findings = [finding]
    result.checks_run = 10
    result.checks_passed = 9
    result.overall_score = 90.0
    result.pillar_scores = {
        "security": PillarScore(pillar="security", score=90.0, findings_count=1, high=1),
        "reliability": PillarScore(pillar="reliability", score=100.0),
        "cost_optimization": PillarScore(pillar="cost_optimization", score=100.0),
        "operational_excellence": PillarScore(pillar="operational_excellence", score=100.0),
        "performance": PillarScore(pillar="performance", score=100.0),
    }
    return result


@pytest.fixture
def reporter(tmp_path):
    return ExcelReporter(output_dir=tmp_path)


def test_generate_creates_xlsx_file(reporter, sample_discovery, sample_assessment):
    output_path = reporter.generate(sample_discovery, sample_assessment, scope_label="test-sub")

    assert output_path.exists()
    assert output_path.suffix == ".xlsx"


def test_dashboard_sheet_exists(reporter, sample_discovery, sample_assessment):
    output_path = reporter.generate(sample_discovery, sample_assessment, scope_label="test-sub")

    workbook = load_workbook(output_path)
    sheet = workbook["Dashboard"]

    assert workbook.sheetnames[0] == "Dashboard"
    assert sheet["A1"].value == "WARA Assessment Dashboard — test-sub"


def test_dashboard_sheet_has_charts(reporter, sample_discovery, sample_assessment):
    output_path = reporter.generate(sample_discovery, sample_assessment, scope_label="test-sub")

    workbook = load_workbook(output_path)
    sheet = workbook["Dashboard"]

    assert len(sheet._charts) == 2


def test_recommendations_sheet_grouped(reporter, sample_discovery, sample_assessment):
    sample_assessment.findings = [
        Finding(
            rule_id="SEC-001",
            title="TLS check",
            pillar="security",
            caf_area="security",
            alz_area="security",
            severity="high",
            confidence="high",
            recommendation="Fix TLS",
            evidence=[{"id": "vm1", "type": "microsoft.compute/virtualmachines"}],
            references=["https://learn.microsoft.com/tls"],
            remediation_steps=["Step 1"],
        ),
        Finding(
            rule_id="SEC-001",
            title="TLS check",
            pillar="security",
            caf_area="security",
            alz_area="security",
            severity="high",
            confidence="high",
            recommendation="Fix TLS",
            evidence=[{"id": "vm2", "type": "microsoft.compute/virtualmachines"}],
            references=["https://learn.microsoft.com/tls"],
            remediation_steps=["Step 2"],
        ),
        Finding(
            rule_id="REL-001",
            title="Backup check",
            pillar="reliability",
            caf_area="management",
            alz_area="backup",
            severity="medium",
            confidence="medium",
            recommendation="Enable backups",
            evidence=[{"id": "sa1", "type": "microsoft.storage/storageaccounts"}],
            references=["https://learn.microsoft.com/backup"],
            remediation_steps=["Step A"],
        ),
    ]

    output_path = reporter.generate(sample_discovery, sample_assessment, scope_label="test-sub")

    workbook = load_workbook(output_path)
    sheet = workbook["Recommendations"]
    rows = [
        [sheet[f"A{row}"].value, sheet[f"D{row}"].value, sheet[f"G{row}"].value]
        for row in range(2, sheet.max_row + 1)
    ]

    assert len(rows) == 2
    assert rows[0] == ["high", 2, "SEC-001"]
    assert rows[1] == ["medium", 1, "REL-001"]


def test_findings_detail_has_all_findings(reporter, sample_discovery, sample_assessment):
    extra_finding = Finding(
        rule_id="REL-001",
        title="Backup check",
        pillar="reliability",
        caf_area="management",
        alz_area="backup",
        severity="medium",
        confidence="medium",
        recommendation="Enable backups",
        remediation_steps=["Step A"],
    )
    sample_assessment.findings.append(extra_finding)

    output_path = reporter.generate(sample_discovery, sample_assessment, scope_label="test-sub")

    workbook = load_workbook(output_path)
    sheet = workbook["Findings Detail"]

    assert sheet.max_row - 1 == len(sample_assessment.findings)


def test_resource_inventory_populated(reporter, sample_discovery, sample_assessment):
    output_path = reporter.generate(sample_discovery, sample_assessment, scope_label="test-sub")

    workbook = load_workbook(output_path)
    sheet = workbook["Resource Inventory"]
    resource_types = [sheet[f"A{row}"].value for row in range(2, sheet.max_row + 1)]

    assert "microsoft.compute/virtualmachines" in resource_types
    assert "microsoft.storage/storageaccounts" in resource_types


def test_pillar_scores_sheet(reporter, sample_discovery, sample_assessment):
    output_path = reporter.generate(sample_discovery, sample_assessment, scope_label="test-sub")

    workbook = load_workbook(output_path)
    sheet = workbook["Pillar Scores"]
    pillar_rows = [sheet[f"A{row}"].value for row in range(2, 7)]

    assert pillar_rows == [
        "Security",
        "Reliability",
        "Cost Optimization",
        "Operational Excellence",
        "Performance Efficiency",
    ]
    assert sheet["A7"].value == "Average / Total"


def test_remediation_roadmap_sorted_by_severity(reporter, sample_discovery, sample_assessment):
    sample_assessment.findings = [
        Finding(
            rule_id="LOW-001",
            title="Low item",
            pillar="performance",
            caf_area="management",
            alz_area="performance",
            severity="low",
            confidence="low",
            recommendation="Low fix",
            remediation_steps=["Low step"],
        ),
        Finding(
            rule_id="CRIT-001",
            title="Critical item",
            pillar="security",
            caf_area="security",
            alz_area="security",
            severity="critical",
            confidence="high",
            recommendation="Critical fix",
            remediation_steps=["Critical step"],
        ),
        Finding(
            rule_id="HIGH-001",
            title="High item",
            pillar="reliability",
            caf_area="management",
            alz_area="reliability",
            severity="high",
            confidence="high",
            recommendation="High fix",
            remediation_steps=["High step"],
        ),
    ]

    output_path = reporter.generate(sample_discovery, sample_assessment, scope_label="test-sub")

    workbook = load_workbook(output_path)
    sheet = workbook["Remediation Roadmap"]
    severities = [sheet[f"D{row}"].value for row in range(2, sheet.max_row + 1)]

    assert severities == ["critical", "high", "low"]


def test_summary_statistics_sheet(reporter, sample_discovery, sample_assessment):
    output_path = reporter.generate(sample_discovery, sample_assessment, scope_label="test-sub")

    workbook = load_workbook(output_path)
    sheet = workbook["Summary Statistics"]
    stats = {sheet[f"A{row}"].value: sheet[f"B{row}"].value for row in range(2, sheet.max_row + 1)}

    assert stats["Scope"] == "test-sub"
    assert stats["Scope Type"] == DiscoveryScope.SUBSCRIPTION.value
    assert stats["Total Resources"] == 7
    assert stats["Total Findings"] == 1
    assert stats["Checks Run"] == 10
    assert stats["Checks Passed"] == 9
    assert stats["Overall Score"] == 90


def test_empty_assessment_still_generates(reporter, sample_discovery):
    assessment = AssessmentResult(scope="test-sub")
    output_path = reporter.generate(sample_discovery, assessment, scope_label="test-sub")

    workbook = load_workbook(output_path)

    assert output_path.exists()
    assert workbook.sheetnames == [
        "Dashboard",
        "Recommendations",
        "Findings Detail",
        "Resource Inventory",
        "Pillar Scores",
        "Remediation Roadmap",
        "Summary Statistics",
    ]
